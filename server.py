import wsgiref.simple_server
import openpyxl
from openpyxl import Workbook
import os
import cgi

import pages
import DB
from CONFIG import DATABASE_PATH


def application(environ, start_response):
    path = environ["PATH_INFO"]
    method = environ["REQUEST_METHOD"]
    field_storage = cgi.FieldStorage(
        fp=environ['wsgi.input'],
        environ=environ,
        keep_blank_values=True
    )
    if path.endswith('.css'):
        content_type = "text/css"
        with open(path[1:], 'r') as file:
            response = file.read().encode()
        status = '200 OK'
    else:
        content_type = "text/html"
        if path == "/":
            pass
        elif path == "/users/":
            if method == "POST":
                second_name = field_storage.getvalue('second_name')
                first_name = field_storage.getvalue('first_name')
                patronymic = field_storage.getvalue('patronymic')
                region = field_storage.getvalue('region')
                city = field_storage.getvalue('city')
                phone = field_storage.getvalue('phone')
                email = field_storage.getvalue('email')
                data = ((second_name, first_name, patronymic, region, city, phone, email),)
                DB.insert_user(data)

            index = pages.show_users()
            response = index.encode()
            status = "200 OK"
        elif path == "/users/add/":

            index = pages.add_user()
            response = index.encode()
            status = "200 OK"
        elif path == "/users/export/xlsx/":
            if method == "POST":
                fileitem = field_storage['file']
                if fileitem.filename:
                    fn = os.path.basename(fileitem.filename)
                    file_path = 'files/' + fn
                    open(file_path, 'wb').write(fileitem.file.read())
                wb = openpyxl.load_workbook('files/users.xlsx')
                ws = wb.active
                data = []
                for row in ws.iter_rows(min_row=2, min_col=2, values_only=True):
                    line = []
                    for item in row:
                        line.append(item)
                    data.append(line)
                for row in data:
                    if row[3] != None:
                        row[3] = DB.get_region_id(row[3])[0]
                    if row[4] != None:
                        row[4] = DB.get_city_id(row[4])[0]
                DB.insert_user(data)

            index = pages.export_xlsx()
            response = index.encode()
            status = "200 OK"
        elif path == "/users/downoload/xlsx/":
            wb = Workbook()

            ws = wb.active

            column_names = ['id', 'second_name', 'first_name', 'patronymic', 'region_name', 'city_name', 'phone', 'email']
            users_data = DB.get_users(region_city_names=True)

            ws.append(column_names)
            for row in users_data:
                ws.append(row)

            wb.save("users.xlsx")
            file_url = 'users.xlsx'
            headers = [('Content-Description', 'File Transfer'),
                       ('Content-Type', 'application/octet-stream'),
                       ('Content-Disposition', 'attachement; filename="' + os.path.basename(file_url) + '"'),
                       ('Expires', '0'),
                       ('Cache-Control', 'must-revalidate'),
                       ('Pragma', 'public'),
                       ('Content-Length', str(os.stat(file_url).st_size))]

            file_download = open(file_url, 'rb')
            status = "200 OK"
            start_response(status, headers)
            return wsgiref.util.FileWrapper(file_download)
        else:
            response = b"<h1>Not found</h1><p>Entered path not found</p>"
            status = "404 Not Found"


    headers = [
        ("Content-Type", content_type),
        ("Content-Length", str(len(response)))
    ]
    start_response(status, headers)

    return [response]


if __name__ == "__main__":
    if not os.path.exists(DATABASE_PATH):
        with open(DATABASE_PATH, 'w'):
            DB.create_tables()
    w_s = wsgiref.simple_server.make_server(
        host="localhost",
        port=8080,
        app=application
    )
    w_s.serve_forever()